# ============================================================
# LLM + MCDA (TOPSIS + PROMETHEE II)
# Backend: Python + Flask
#
## Kaj dela backend?
# 1) Vrne seznam Ollama modelov (iz /api/tags).
# 2) Ustvari povzetek (Ollama /api/generate, keep_alive=0).
# 3) Izmeri čas + RAM peak/avg (delta v MB) za OLLAMA proces.
# 4) Izvede MCDA metode (TOPSIS + PROMETHEE I).
# 5) Shrani zgodovino testov v JSON (lokalno).
# 6) Izvozi Excel (XLSX) z več sheeti.
# ============================================================

import flask
import requests
import psutil
import time
import threading
import os
import json
import math
import statistics
import openpyxl

# ------------------------------------------------------------
# Osnovne nastavitve
# ------------------------------------------------------------

# URL lokalnega Ollama strežnika.
# Privzeto: http://127.0.0.1:11434
OLLAMA_URL = "http://127.0.0.1:11434"

# Kje hranimo zgodovino testov (JSON datoteka).
HISTORY_FILE = os.path.join(os.path.dirname(__file__), "..", "data", "history.json")

# Kriteriji za MCDA (ključ, prikazno ime, tip: cost=MIN, benefit=MAX)
CRITERIA = [
    {"key": "time_s", "name": "Čas (s)", "type": "cost"},
    {"key": "ram_peak_mb", "name": "RAM peak (MB)", "type": "cost"},
    {"key": "ram_avg_mb", "name": "RAM avg (MB)", "type": "cost"},
    {"key": "output_pct", "name": "Output (%)", "type": "cost"},
    {"key": "quality", "name": "Kvaliteta povzetka (1–5)", "type": "benefit"},
    {"key": "hallucinations", "name": "Količina halucinacij (1–5)", "type": "cost"},
]

# Flask aplikacija.
# static_folder kaže na mapo "frontend", da lahko serviramo HTML/CSS/JS.
app = flask.Flask(__name__, static_folder="../frontend", static_url_path="")


# ------------------------------------------------------------
# Zgodovina (JSON)
# ------------------------------------------------------------

def load_history():
    """Prebere zgodovino testov iz JSON datoteke (če obstaja)."""
    try:
        if not os.path.exists(HISTORY_FILE):
            return []
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return data
        return []
    except Exception:
        # Če je datoteka poškodovana, raje vrnemo prazno zgodovino.
        return []


def save_history(tests):
    """Shrani seznam testov v JSON datoteko."""
    try:
        os.makedirs(os.path.dirname(HISTORY_FILE), exist_ok=True)
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(tests, f, ensure_ascii=False, indent=2)
    except Exception:
        # Če shranjevanje ne uspe, aplikacija naj vseeno dela.
        pass


def clear_history():
    """Izbriše vso zgodovino testov."""
    save_history([])


# ------------------------------------------------------------
# Merjenje RAM (Ollama proces)
# ------------------------------------------------------------

def get_ollama_rss_total():
    """
    Vrne trenutni RSS (RAM) za Ollama proces(e) v bajtih.

    Zakaj to rabimo?
    - Flask aplikacija teče v svojem procesu, Ollama pa v svojem.
    - Če merimo RAM samo od Flask-a, dobimo skoraj 0 MB.
    - Zato poiščemo proces(e), katerih ime vsebuje 'ollama', in seštejemo njihov RSS.

    Opomba:
    - Na Windows se lahko proces imenuje 'ollama.exe', na Linux/macOS pa 'ollama'.
    """
    total = 0
    try:
        for p in psutil.process_iter(["name", "memory_info"]):
            name = (p.info.get("name") or "").lower()
            if "ollama" in name:
                mi = p.info.get("memory_info")
                if mi:
                    total += mi.rss
    except Exception:
        return 0
    return total


def measure_ram_while_running(stop_flag, samples, interval_seconds):
    """
    To teče v ločeni niti (thread).

    Medtem ko stop_flag ni True:
    - vsake interval_seconds izmeri RAM (RSS) OLLAMA procesa
    - shranjuje meritve v seznam 'samples'

    Tako dobimo RAM peak in RAM average med izvajanjem povzetka.
    """
    while True:
        if stop_flag["stop"]:
            break
        try:
            samples.append(get_ollama_rss_total())
        except Exception:
            pass
        time.sleep(interval_seconds)


# ------------------------------------------------------------
# Klic Ollama (povzetek)
# ------------------------------------------------------------

def call_ollama_summarize(model, input_text):
    """Pokliče Ollama /api/generate in vrne (summary, token_in, token_out)."""
    prompt = (
        "You are a careful summarization assistant.\n\n"
        "TASK:\n"
        "Summarize the text below in ENGLISH.\n\n"
        "PRIORITIES (in this order):\n"
        "1) Be as short as possible while still capturing the core message.\n"
        "2) Be strictly faithful: do NOT add any facts, numbers, examples, or causes not explicitly stated.\n"
        "3) Keep the summary clear, coherent, and neutral.\n\n"
        "RULES:\n"
        "- One paragraph only (no bullets, no headings).\n"
        "- If you are unsure whether something is in the text, leave it out.\n\n"
        "TEXT:\n" + input_text
    )

    payload = {
        "model": model,
        "prompt": prompt,
        "stream": False,
        "keep_alive": 0  # pomembno: po koncu naj se model sprosti iz RAM-a
    }

    r = requests.post(OLLAMA_URL + "/api/generate", json=payload, timeout=600)
    r.raise_for_status()
    data = r.json()

    summary = data.get("response", "") or ""
    token_in = data.get("prompt_eval_count", None)
    token_out = data.get("eval_count", None)

    return summary, token_in, token_out


# ------------------------------------------------------------
# MCDA: TOPSIS
# ------------------------------------------------------------
# Ideja TOPSIS:
# 1) Vse kriterije spravimo v "več je bolje" (benefit).
#    - V tej implementaciji to naredimo tako, da COST kriterije pomnožimo z -1.
# 2) Normaliziramo (vektorska normalizacija), da kriteriji z velikimi številkami
#    (npr. RAM v MB) ne "povozijo" kriterijev z majhnimi (npr. kvaliteta 1..3).
# 3) Upoštevamo uteži (w_j): pomembnejši kriterij bolj vpliva.
# 4) Določimo idealno (najboljše) in anti-idealno (najslabše) rešitev.
# 5) Izračunamo razdaljo vsake alternative do ideala (D+) in anti-ideala (D-).
# 6) Izračunamo bližino ideala C* = D- / (D+ + D-).
#    - Večji C* pomeni bližje idealu => boljša alternativa.

def topsis_vector_norm(matrix):
    """
    1. korak TOPSIS: vektorska normalizacija.
    Formula:
        r_ij = x_ij / sqrt(sum_i x_ij^2)

    - Za vsak stolpec (kriterij) izračunamo vsoto kvadratov.
    - Nato vsak element delimo s korenom te vsote.
    - Rezultat je brez enot in primerljiv med kriteriji.
    """
    m = len(matrix)        # število alternativ
    n = len(matrix[0])     # število kriterijev

    # col_sums[j] = sum_i x_ij^2  (vsota kvadratov v stolpcu j)
    col_sums = [0.0] * n
    for j in range(n):
        s = 0.0
        for i in range(m):
            s += float(matrix[i][j]) ** 2
        col_sums[j] = s

    # norm[i][j] = r_ij
    norm = []
    for i in range(m):
        row = []
        for j in range(n):
            # denom = sqrt(sum_i x_ij^2); če je 0, uporabimo 1 da se izognemo deljenju z 0
            denom = math.sqrt(col_sums[j]) if col_sums[j] != 0 else 1.0
            row.append(round(float(matrix[i][j]) / denom, 6))
        norm.append(row)

    return norm


def topsis_weighted(norm_matrix, weights):
    """
    2. korak TOPSIS: utežena normalizacija.
    Formula:
        v_ij = w_j * r_ij

    - r_ij je normalizirana vrednost kriterija.
    - w_j predstavlja pomembnost kriterija.
    - Po tem koraku dobimo uteženo normalizirano matriko V.
    """
    m = len(norm_matrix)
    n = len(norm_matrix[0])

    out = []
    for i in range(m):
        row = []
        for j in range(n):
            row.append(round(float(norm_matrix[i][j]) * float(weights[j]), 6))
        out.append(row)
    return out


def topsis_ideal_antiideal(weighted_matrix):
    """
    3. korak TOPSIS: določitev idealne (A*) in antiidealne (A-) točke.

    Ker smo COST kriterije že pretvorili v benefit (z negacijo),
    lahko za vsak kriterij j:
        ideal[j] = max_i v_ij   (najboljša vrednost)
        anti[j]  = min_i v_ij   (najslabša vrednost)
    """
    m = len(weighted_matrix)
    n = len(weighted_matrix[0])

    ideal = []
    anti = []
    for j in range(n):
        col = [weighted_matrix[i][j] for i in range(m)]
        ideal.append(max(col))
        anti.append(min(col))
    return ideal, anti


def topsis_distances(weighted_matrix, ideal, anti):
    """
    4. korak TOPSIS: razdalje do ideala in anti-ideala.
    Uporabimo evklidsko razdaljo:

        D+(i) = sqrt( sum_j (v_ij - A*_j)^2 )
        D-(i) = sqrt( sum_j (v_ij - A-_j)^2 )

    - D+ manjše => alternativa je bližje idealu (dobro)
    - D- večje => alternativa je dlje od najslabše točke (dobro)
    """
    m = len(weighted_matrix)
    n = len(weighted_matrix[0])

    d_plus = []
    d_minus = []
    for i in range(m):
        s1 = 0.0
        s2 = 0.0
        for j in range(n):
            s1 += (weighted_matrix[i][j] - ideal[j]) ** 2
            s2 += (weighted_matrix[i][j] - anti[j]) ** 2
        d_plus.append(math.sqrt(s1))
        d_minus.append(math.sqrt(s2))
    return d_plus, d_minus


def topsis_closeness(d_plus, d_minus):
    """
    5. korak TOPSIS: koeficient bližine ideala (C*).

        C*(i) = D-(i) / (D+(i) + D-(i))

    - Če je alternativa zelo blizu ideala: D+ ~ 0 => C* ~ 1.
    - Če je alternativa blizu anti-ideala: D- ~ 0 => C* ~ 0.
    """
    m = len(d_plus)
    c = []
    for i in range(m):
        denom = d_plus[i] + d_minus[i]
        c.append((d_minus[i] / denom) if denom != 0 else 0.0)
    return c


def run_topsis(decision_matrix, weights, criteria_types, alternatives):
    """
    Glavna funkcija TOPSIS:
    Vhod:
      - decision_matrix: matrika alternativ x kriterijev (raw vrednosti)
      - weights: uteži kriterijev (v istem vrstnem redu kot stolpci)
      - criteria_types: seznam "cost" ali "benefit" za vsak kriterij
      - alternatives: imena alternativ (vrstic)

    Izhod:
      - seznam slovarjev: Alternative, D+, D-, C*, Rank
    """

    # (0) Pretvorba COST -> BENEFIT z negacijo.
    # Zakaj? TOPSIS v nadaljevanju uporablja "ideal = max".
    # Če je kriterij cost (manj je bolje), ga obrnemo tako, da večja (manj negativna)
    # vrednost pomeni boljši rezultat.
    adjusted = []
    for row in decision_matrix:
        r = []
        for j, v in enumerate(row):
            if criteria_types[j] == "cost":
                r.append(float(v) * -1.0)
            else:
                r.append(float(v))
        adjusted.append(r)

    # (1) normalizacija
    norm_m = topsis_vector_norm(adjusted)

    # (2) uteževanje
    weighted = topsis_weighted(norm_m, weights)

    # (3) ideal in anti-ideal
    ideal, anti = topsis_ideal_antiideal(weighted)

    # (4) razdalje
    d_plus, d_minus = topsis_distances(weighted, ideal, anti)

    # (5) C*
    c_star = topsis_closeness(d_plus, d_minus)

    # Rang: večji C* je boljše
    order = sorted(range(len(c_star)), key=lambda i: c_star[i], reverse=True)
    ranks = [0] * len(c_star)
    for r, idx in enumerate(order, start=1):
        ranks[idx] = r

    rows = []
    for i, alt in enumerate(alternatives):
        rows.append({
            "Alternative": alt,
            "D_plus": round(d_plus[i], 6),
            "D_minus": round(d_minus[i], 6),
            "C_star": round(c_star[i], 6),
            "Rank": ranks[i]
        })

    return rows


# ------------------------------------------------------------
# MCDA: PROMETHEE II
# ------------------------------------------------------------
# Ideja PROMETHEE II:
# 1) Normalizacija (min-max) -> vse kriterije spravimo v 0..1 in "več je bolje".
# 2) Parno primerjamo alternative (a proti b) po vsakem kriteriju:
#      diff = v_a - v_b
#      P(diff) = max(diff, 0)   (t.i. "usual" / V-shape brez praga)
#    To pomeni:
#      - če je a boljša od b -> pozitivna preferenca
#      - če je slabša ali enaka -> 0 (ni preference)
# 3) Utežimo preference po kriterijih in jih seštejemo:
#      π(a,b) = sum_j w_j * P_j(a,b)
# 4) Iz π(a,b) izračunamo tokove:
#      φ+(a) = povprečna preferenca a nad drugimi (odhodni tok)
#      φ-(a) = povprečna preferenca drugih nad a (dohodni tok)
#      φ(a)  = φ+(a) - φ-(a)    (neto tok; večji je boljše)
# 5) Rang: sort po φ padajoče.

def promethee_normalize(matrix, criteria_types):
    """
    1. korak PROMETHEE: min-max normalizacija po stolpcih.

    - Če je kriterij cost:
        norm = (max - x) / (max - min)   (manjše x -> večji norm -> bolje)
    - Če je kriterij benefit:
        norm = (x - min) / (max - min)   (večje x -> večji norm -> bolje)

    Rezultat: norm[i][j] je vedno v [0,1] in pomeni "več je bolje".
    """
    m = len(matrix)
    n = len(matrix[0])

    # priprava matrike
    norm = [[0.0 for _ in range(n)] for _ in range(m)]

    for j in range(n):
        col = [float(matrix[i][j]) for i in range(m)]
        col_min = min(col)
        col_max = max(col)

        # Če so vse vrednosti enake, kriterij ne loči alternativ -> vse 0
        if col_max == col_min:
            for i in range(m):
                norm[i][j] = 0.0
            continue

        for i in range(m):
            x = float(matrix[i][j])
            if criteria_types[j] == "cost":
                # manj je bolje -> obrnemo os
                norm[i][j] = (col_max - x) / (col_max - col_min)
            else:
                # več je bolje
                norm[i][j] = (x - col_min) / (col_max - col_min)

    return norm


def run_promethee(decision_matrix, weights, criteria_types, alternatives):
    """
    Glavna funkcija PROMETHEE II.
    Vračamo: Alternative, Phi+, Phi-, Phi, Rank
    """

    # (0) Normalizacija uteži: PROMETHEE pričakuje, da uteži predstavljajo deleže.
    # Če vsota ni 1, jih tu delimo z vsoto.
    wsum = sum(weights) if sum(weights) != 0 else 1.0
    w = [float(x) / wsum for x in weights]

    # (1) Normaliziramo matriko v 0..1, "več je bolje"
    norm = promethee_normalize(decision_matrix, criteria_types)
    m = len(norm)      # alternative
    n = len(norm[0])   # kriteriji

    # (2) agg[a][b] bo predstavljal agregirano preferenco π(a,b)
    #     torej: koliko alternativa a preferira alternativo b (uteženo)
    agg = [[0.0 for _ in range(m)] for _ in range(m)]

    # (3) Izračun π(a,b) za vse pare alternativ
    for a in range(m):
        for b in range(m):
            if a == b:
                agg[a][b] = 0.0
                continue

            s = 0.0
            for j in range(n):
                # diff > 0 pomeni, da je a boljša od b pri kriteriju j
                diff = norm[a][j] - norm[b][j]

                # "usual" preferenčna funkcija:
                # P = diff, če je diff > 0, sicer 0
                pref = diff if diff > 0 else 0.0

                # utežen prispevek kriterija j k agregirani preferenci
                s += pref * w[j]

            agg[a][b] = s

    # (4) Izračun pozitivnega toka φ+(a):
    # povprečje preferenc a nad vsemi drugimi b
    phi_plus = []
    for a in range(m):
        phi_plus.append(sum(agg[a]) / (m - 1) if m > 1 else 0.0)

    # (5) Izračun negativnega toka φ-(a):
    # povprečje preferenc vseh drugih b nad a
    phi_minus = []
    for a in range(m):
        col_sum = 0.0
        for b in range(m):
            col_sum += agg[b][a]
        phi_minus.append(col_sum / (m - 1) if m > 1 else 0.0)

    # (6) Neto tok φ(a) = φ+(a) - φ-(a)
    # večji φ pomeni, da alternativa v povprečju bolj premaga druge kot izgublja proti njim
    phi = [phi_plus[i] - phi_minus[i] for i in range(m)]

    # Rang: večji φ je boljše
    order = sorted(range(len(phi)), key=lambda i: phi[i], reverse=True)
    ranks = [0] * len(phi)
    for r, idx in enumerate(order, start=1):
        ranks[idx] = r

    rows = []
    for i, alt in enumerate(alternatives):
        rows.append({
            "Alternative": alt,
            "Phi_plus": round(phi_plus[i], 6),
            "Phi_minus": round(phi_minus[i], 6),
            "Phi": round(phi[i], 6),
            "Rank": ranks[i]
        })

    return rows



# ------------------------------------------------------------
# Excel izvoz (XLSX)
# ------------------------------------------------------------

def build_xlsx_bytes(all_tests, current_test, mode):
    """Naredi Excel v pomnilniku in vrne bytes."""
    wb = openpyxl.Workbook()

    ws_tests = wb.active
    ws_tests.title = "Tests"

    ws_matrix = wb.create_sheet("MCDA_Matrix")
    ws_weights = wb.create_sheet("Weights")
    ws_topsis = wb.create_sheet("TOPSIS")
    ws_prom = wb.create_sheet("PROMETHEE")

    def safe_add_row(ws, row):
        ws.append(row)

    tests_to_export = []
    if mode == "current":
        if current_test:
            tests_to_export.append(current_test)
    else:
        tests_to_export.extend(all_tests)
        if current_test:
            tests_to_export.append(current_test)

    # --------------------------------------------------------
    # Sheet: Tests
    # (1 vrstica = 1 alternativa v testu)
    # --------------------------------------------------------
    safe_add_row(ws_tests, [
        "TestID", "Alternative",
        "Time(s)", "RAM peak (MB)", "RAM avg (MB)",
        "Output(%)", "Input chars", "Output chars",
        "Token in", "Token out", "Quality(1-5)", "Hallucinations(1-5)",
        "Input text", "Summary"
    ])

    for t in tests_to_export:
        test_id = t.get("test_id", "")
        input_chars = t.get("input_chars", "")

        for alt in t.get("alternatives", []):
            safe_add_row(ws_tests, [
                test_id,
                alt.get("alternative", ""),
                alt.get("time_s", ""),
                alt.get("ram_peak_mb", ""),
                alt.get("ram_avg_mb", ""),
                alt.get("output_pct", ""),
                input_chars,
                alt.get("output_chars", ""),
                alt.get("token_in", ""),
                alt.get("token_out", ""),
                alt.get("quality", ""),
                alt.get("hallucinations", ""),
                (t.get("input_text", "") or alt.get("input_text", ""))[:2000],
                alt.get("summary", "")[:1000]  # da Excel ne postane ogromen
            ])

    # --------------------------------------------------------
    # Sheet: Weights
    # --------------------------------------------------------
    safe_add_row(ws_weights, ["TestID", "Criterion", "Weight"])
    for t in tests_to_export:
        test_id = t.get("test_id", "")
        w = t.get("weights", {})
        for c in CRITERIA:
            safe_add_row(ws_weights, [test_id, c["name"], w.get(c["key"], "")])

    # --------------------------------------------------------
    # Sheet: MCDA_Matrix
    # --------------------------------------------------------
    safe_add_row(ws_matrix, ["TestID", "Alternative"] + [c["name"] for c in CRITERIA])
    for t in tests_to_export:
        test_id = t.get("test_id", "")
        for alt in t.get("alternatives", []):
            safe_add_row(ws_matrix, [
                test_id,
                alt.get("alternative", ""),
                alt.get("time_s", ""),
                alt.get("ram_peak_mb", ""),
                alt.get("ram_avg_mb", ""),
                alt.get("output_pct", ""),
                alt.get("quality", ""),
                alt.get("hallucinations", "")
            ])

    # --------------------------------------------------------
    # Sheet: TOPSIS
    # --------------------------------------------------------
    safe_add_row(ws_topsis, ["TestID", "Alternative", "D+", "D-", "C*", "Rank"])
    for t in tests_to_export:
        test_id = t.get("test_id", "")
        for r in t.get("topsis", []):
            safe_add_row(ws_topsis, [
                test_id,
                r.get("Alternative", ""),
                r.get("D_plus", ""),
                r.get("D_minus", ""),
                r.get("C_star", ""),
                r.get("Rank", "")
            ])

    # --------------------------------------------------------
    # Sheet: PROMETHEE
    # --------------------------------------------------------
    safe_add_row(ws_prom, ["TestID", "Alternative", "Phi+", "Phi-", "Phi", "Rank"])
    for t in tests_to_export:
        test_id = t.get("test_id", "")
        for r in t.get("promethee", []):
            safe_add_row(ws_prom, [
                test_id,
                r.get("Alternative", ""),
                r.get("Phi_plus", ""),
                r.get("Phi_minus", ""),
                r.get("Phi", ""),
                r.get("Rank", "")
            ])

    import io
    buff = io.BytesIO()
    wb.save(buff)
    return buff.getvalue()


# ------------------------------------------------------------
# ROUTES: Frontend
# ------------------------------------------------------------

@app.route("/")
def index():
    return app.send_static_file("index.html")


# ------------------------------------------------------------
# ROUTES: API
# ------------------------------------------------------------

@app.route("/api/models", methods=["GET"])
def api_models():
    """Vrne seznam modelov iz Ollama (/api/tags)."""
    try:
        r = requests.get(OLLAMA_URL + "/api/tags", timeout=10)
        r.raise_for_status()
        data = r.json()

        models = []
        for m in data.get("models", []):
            name = m.get("name", "")
            if name:
                models.append(name)

        return flask.jsonify({"ok": True, "models": models})
    except Exception as e:
        return flask.jsonify({
            "ok": False,
            "error": "Ollama ni dosegljiv. Preveri, če Ollama teče. (" + str(e) + ")"
        }), 500


@app.route("/api/summarize", methods=["POST"])
def api_summarize():
    """Ustvari povzetek + meritve. Ne shranjuje v zgodovino."""
    try:
        body = flask.request.get_json(force=True)
        model = (body.get("model") or "").strip()
        input_text = (body.get("input_text") or "").strip()

        if model == "" or input_text == "":
            return flask.jsonify({"ok": False, "error": "Manjka model ali vhodno besedilo."}), 400

        # Baseline RAM OLLAMA procesa
        baseline = get_ollama_rss_total()

        # Začnemo meriti RAM v ločeni niti
        stop_flag = {"stop": False}
        samples = []
        t = threading.Thread(target=measure_ram_while_running, args=(stop_flag, samples, 0.05))
        t.daemon = True
        t.start()

        # Čas izvajanja
        t0 = time.perf_counter()
        summary, token_in, token_out = call_ollama_summarize(model, input_text)
        t1 = time.perf_counter()

        # Ustavimo merjenje RAM
        stop_flag["stop"] = True
        try:
            t.join(timeout=1.0)
        except Exception:
            pass

        if len(samples) == 0:
            samples = [baseline]

        peak = max(samples)
        avg = statistics.mean(samples)

        # delta v MB
        peak_delta = max(0.0, (peak - baseline) / (1024 * 1024))
        avg_delta = max(0.0, (avg - baseline) / (1024 * 1024))

        input_chars = len(input_text)
        output_chars = len(summary)
        output_pct = (output_chars / input_chars * 100.0) if input_chars > 0 else 0.0

        return flask.jsonify({
            "ok": True,
            "model": model,
            "summary": summary,
            "time_s": round((t1 - t0), 1),
            "ram_peak_mb": round(peak_delta, 3),
            "ram_avg_mb": round(avg_delta, 3),
            "output_pct": round(output_pct, 2),
            "input_chars": input_chars,
            "output_chars": output_chars,
            "token_in": token_in,
            "token_out": token_out
        })

    except requests.exceptions.RequestException as e:
        return flask.jsonify({"ok": False, "error": "Napaka pri klicu Ollama: " + str(e)}), 500
    except Exception as e:
        return flask.jsonify({"ok": False, "error": "Nepričakovana napaka: " + str(e)}), 500


@app.route("/api/mcda", methods=["POST"])
def api_mcda():
    """Izvede TOPSIS in PROMETHEE II. Test se shrani v zgodovino samo ob save_to_history=True."""
    try:
        body = flask.request.get_json(force=True)
        test = body.get("test", {})
        save_to_history_flag = bool(body.get("save_to_history", True))

        alternatives = test.get("alternatives", [])
        weights = test.get("weights", {})

        # Za MCDA potrebujemo vsaj 2 alternativi.
        # Če je samo 1 alternativa, ne moremo narediti primerjave.
        if len(alternatives) < 2:
            return flask.jsonify({"ok": False, "error": "Za izvedbo MCDA sta potrebni najmanj 2 alternativi."}), 400

        if len(alternatives) == 0:
            return flask.jsonify({"ok": False, "error": "Ni alternativ v matriki."}), 400

        # Uteži v pravem vrstnem redu
        w_list = []
        for c in CRITERIA:
            w_list.append(float(weights.get(c["key"], 0.0)))

        w_sum = sum(w_list)
        if w_sum <= 0:
            return flask.jsonify({"ok": False, "error": "Uteži morajo biti > 0 in vsota naj bo 1."}), 400

        # dovolimo majhno napako
        if abs(w_sum - 1.0) > 0.01:
            return flask.jsonify({"ok": False, "error": "Vsota uteži mora biti približno 1. Trenutno: " + str(round(w_sum, 4))}), 400

        decision_matrix = []
        alt_names = []
        for alt in alternatives:
            alt_names.append(alt.get("alternative", ""))
            row = []
            for c in CRITERIA:
                row.append(float(alt.get(c["key"], 0.0)))
            decision_matrix.append(row)

        criteria_types = [c["type"] for c in CRITERIA]

        topsis_rows = run_topsis(decision_matrix, w_list, criteria_types, alt_names)
        prom_rows = run_promethee(decision_matrix, w_list, criteria_types, alt_names)

        test["topsis"] = topsis_rows
        test["promethee"] = prom_rows

        if save_to_history_flag:
            hist = load_history()
            hist.append(test)
            save_history(hist)

        return flask.jsonify({"ok": True, "topsis": topsis_rows, "promethee": prom_rows})

    except Exception as e:
        return flask.jsonify({"ok": False, "error": "Napaka pri MCDA: " + str(e)}), 500


@app.route("/api/history/clear", methods=["POST"])
def api_history_clear():
    clear_history()
    return flask.jsonify({"ok": True})


@app.route("/api/export/current", methods=["POST"])
def api_export_current():
    try:
        body = flask.request.get_json(force=True)
        current_test = body.get("test", None)
        data = build_xlsx_bytes(load_history(), current_test, "current")

        return flask.Response(
            data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=current_test.xlsx"}
        )
    except Exception as e:
        return flask.jsonify({"ok": False, "error": "Napaka pri izvozu: " + str(e)}), 500


@app.route("/api/export/all", methods=["POST"])
def api_export_all():
    try:
        body = flask.request.get_json(force=True)
        current_test = body.get("test", None)
        data = build_xlsx_bytes(load_history(), current_test, "all")

        return flask.Response(
            data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=all_tests.xlsx"}
        )
    except Exception as e:
        return flask.jsonify({"ok": False, "error": "Napaka pri izvozu: " + str(e)}), 500


# ------------------------------------------------------------
# Zagon
# ------------------------------------------------------------
if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
